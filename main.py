import os, re, tempfile
from typing import List
from datetime import datetime

from fastapi import FastAPI, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter

# ── optional imports ────────────────────────────────────────────────────────
try:
    import fitz  # PyMuPDF
    PYMUPDF_OK = True
except ImportError:
    PYMUPDF_OK = False

try:
    from pdf2image import convert_from_path
    import pytesseract
    OCR_OK = True
except ImportError:
    OCR_OK = False

# ── app setup ───────────────────────────────────────────────────────────────
app = FastAPI(title="C4 Capital FIDC API")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], allow_methods=["*"], allow_headers=["*"]
)

# ── PDF text extraction (PyMuPDF + OCR fallback) ────────────────────────────
def extrair_texto_pdf(caminho: str) -> str:
    txt = ""
    if PYMUPDF_OK:
        try:
            doc = fitz.open(caminho)
            txt = "".join(p.get_text() for p in doc).strip()
            doc.close()
        except Exception:
            pass

    if not txt and OCR_OK:
        try:
            pages = convert_from_path(caminho, dpi=200)
            for pg in pages:
                txt += pytesseract.image_to_string(pg, lang="por+eng") + "\n"
            txt = txt.strip()
        except Exception:
            pass

    return txt

# ── helpers ────────────────────────────────────────────────────────────────
def _limpo(s: str) -> str:
    return re.sub(r"\D", "", s)

def _valor(s: str) -> float:
    try:
        return float(s.replace(".", "").replace(",", "."))
    except Exception:
        return 0.0

# ── NFS-e extraction ─────────────────────────────────────────────────────────
def _extrair_nfse(txt: str) -> dict:
    d = dict(tipo="nfse", cnpj_ced="", razao_ced="",
             cnpj_sac="", razao_sac="",
             num="", dt="", venc="", valor=0.0,
             chave="", cep="", email="", tel="")

    # Chave de acesso
    m = re.search(r"Chave de Acesso da NFS-e\s+([\d\s]+)", txt)
    if m:
        d["chave"] = re.sub(r"\s", "", m.group(1).split("\n")[0].strip())

    # Número + data de emissão
    m = re.search(r"Numero da NFS-e\s+(\d+)\s+(\d{2}/\d{2}/\d{4})", txt)
    if m:
        d["num"] = m.group(1)
        d["dt"]  = m.group(2)

    # Valor líquido
    m = re.search(r"Valor Liqu?ido da NFS-e\s+([\d\.]+,\d{2})", txt)
    if m:
        d["valor"] = _valor(m.group(1))

    # Vencimento
    m = re.search(r"[Dd]ata\s+[Vv]encimento[:\s]+(\d{2}/\d{2}/\d{4})", txt)
    if m:
        d["venc"] = m.group(1)

    # Locate sections
    prest = re.search(r"Prestador do Servi[cç]o", txt, re.I)
    tomd  = re.search(r"TOMADOR DO SERVI", txt, re.I)

    if prest and tomd:
        p_txt = txt[prest.end(): tomd.start()]

        m = re.search(r"Nome / Nome Empresarial\s+(.+)", p_txt)
        if m:
            d["razao_ced"] = m.group(1).strip()

        cnpjs = re.findall(r"\d{2}[\.\s]?\d{3}[\.\s]?\d{3}[/\s]?\d{4}[-\s]?\d{2}", p_txt)
        if cnpjs:
            d["cnpj_ced"] = _limpo(cnpjs[0])

    if tomd:
        t_txt = txt[tomd.end():]
        stop  = re.search(r"INTERMEDIARIO|SERVI[ÇC]O PRESTADO", t_txt, re.I)
        if stop:
            t_txt = t_txt[: stop.start()]

        m = re.search(r"Nome / Nome Empresarial\s+(.+)", t_txt)
        if m:
            d["razao_sac"] = m.group(1).strip()

        cnpjs = re.findall(r"\d{2}[\.\s]?\d{3}[\.\s]?\d{3}[/\s]?\d{4}[-\s]?\d{2}", t_txt)
        if cnpjs:
            d["cnpj_sac"] = _limpo(cnpjs[0])

        m = re.search(r"CEP\s+(\d{5}[-\s]?\d{3})", t_txt)
        if m:
            d["cep"] = _limpo(m.group(1))

        m = re.search(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}", t_txt)
        if m:
            d["email"] = m.group(0)

        m = re.search(r"Telefone\s+(\(?\d{2}\)?\s?\d{4,5}[-\s]?\d{4})", t_txt)
        if m:
            d["tel"] = _limpo(m.group(1))

    return d

# ── Fatura de locação extraction ─────────────────────────────────────────────
def _extrair_fatura(txt: str) -> list:
    cedente = re.search(r"Raz[ãa]o Social[:\s]+(.+)", txt, re.I)
    cnpj_c  = re.search(r"CNPJ[:\s]+([\d.\/\-]+)", txt, re.I)
    sacado  = re.search(r"(?:Locatário|Tomador)[:\s]+(.+)", txt, re.I)
    cnpj_s  = re.findall(r"\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}", txt)
    parcelas = re.findall(
        r"(\d{2}/\d{2}/\d{4})\s+([\d\.]+,\d{2})", txt
    )
    resultado = []
    for i, (dt_venc, val_str) in enumerate(parcelas):
        resultado.append(dict(
            tipo="fatura",
            cnpj_ced=_limpo(cnpj_c.group(1)) if cnpj_c else "",
            razao_ced=cedente.group(1).strip() if cedente else "",
            cnpj_sac=_limpo(cnpj_s[1]) if len(cnpj_s) > 1 else "",
            razao_sac=sacado.group(1).strip() if sacado else "",
            num=f"PARC-{i+1:02d}",
            dt=datetime.today().strftime("%d/%m/%Y"),
            venc=dt_venc,
            valor=_valor(val_str),
            chave="", cep="", email="", tel=""
        ))
    return resultado

# ── Route PDF to correct extractor ───────────────────────────────────────────
def extrair_pdf(caminho: str):
    txt = extrair_texto_pdf(caminho)
    if re.search(r"[Ff]atura|[Ll]oca[çc][ãa]o|[Pp]arcela", txt):
        resultados = _extrair_fatura(txt)
        return resultados if resultados else [_extrair_nfse(txt)]
    return [_extrair_nfse(txt)]

# ── Excel generation (from scratch) ──────────────────────────────────────────
_THIN = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=9)
_TITLE_FONT  = Font(bold=True, size=11, color="1F3864")
_LABEL_FONT  = Font(bold=True, size=9)
_DATA_FONT   = Font(size=9)

def _cell(ws, addr, value, font=None, fill=None, align=None, border=None, num_fmt=None):
    c = ws[addr]
    c.value = value
    if font:   c.font   = font
    if fill:   c.fill   = fill
    if align:  c.alignment = align
    if border: c.border = border
    if num_fmt: c.number_format = num_fmt
    return c

def preencher_bordero(dados_list: list) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bordero"

    # Column widths
    widths = {"A":5,"B":20,"C":34,"D":14,"E":13,"F":13,"G":26,"H":16,"I":14,"J":46,"K":12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Row 1 – title bar
    ws.merge_cells("A1:F1")
    _cell(ws, "A1", "BORDERÔ DE DESCONTO DE TÍTULOS DE CRÉDITO",
          font=_TITLE_FONT,
          align=Alignment(horizontal="left", vertical="center"))
    ws.row_dimensions[1].height = 20
    _cell(ws, "G1", "Data:", font=_LABEL_FONT,
          align=Alignment(horizontal="right", vertical="center"))
    _cell(ws, "H1", datetime.today().strftime("%d/%m/%Y"),
          font=_DATA_FONT,
          align=Alignment(horizontal="left", vertical="center"))

    # Rows 3-6 – cedente + config
    primeiro = dados_list[0] if dados_list else {}
    _cell(ws, "A3", "CNPJ Cedente:",  font=_LABEL_FONT)
    _cell(ws, "C3", primeiro.get("cnpj_ced",""), font=_DATA_FONT)
    _cell(ws, "D3", "Tipo Oper:",     font=_LABEL_FONT)
    _cell(ws, "E3", "Comissária",     font=_DATA_FONT)
    _cell(ws, "F3", "Tipo NF:",       font=_LABEL_FONT)
    _cell(ws, "G3", "Service",        font=_DATA_FONT)

    _cell(ws, "A4", "Razão Social Cedente:", font=_LABEL_FONT)
    _cell(ws, "C4", primeiro.get("razao_ced",""), font=_DATA_FONT)

    _cell(ws, "A5", "GN:", font=_LABEL_FONT)
    _cell(ws, "C5", "Gustavo Oliveira", font=_DATA_FONT)

    n = len(dados_list)
    last_row = 8 + n
    _cell(ws, "A6", "Valor total do Borderô:", font=Font(bold=True, size=9))
    ws["C6"] = f"=SUM(I9:I{last_row})"
    ws["C6"].font  = Font(bold=True, size=9)
    ws["C6"].number_format = '#,##0.00'

    # Row 8 – column headers
    headers = ["QTD","CPF/CNPJ SACADO","RAZÃO SOCIAL DO SACADO",
               "N.TÍTULO","DT EMISSÃO","VENCIMENTO",
               "E-MAIL","TELEFONE","VALOR R$","CHAVE XML","CEP SACADO"]
    for col_i, h in enumerate(headers, 1):
        c = ws.cell(row=8, column=col_i, value=h)
        c.font   = _HEADER_FONT
        c.fill   = _HEADER_FILL
        c.border = _BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[8].height = 28

    # Data rows
    for i, d in enumerate(dados_list):
        r = 9 + i
        row_data = [
            f"{i+1:02d}",
            d.get("cnpj_sac",""),
            d.get("razao_sac",""),
            d.get("num",""),
            d.get("dt",""),
            d.get("venc",""),
            d.get("email",""),
            d.get("tel",""),
            d.get("valor", 0),
            d.get("chave",""),
            d.get("cep",""),
        ]
        for col_i, val in enumerate(row_data, 1):
            c = ws.cell(row=r, column=col_i, value=val)
            c.font   = _DATA_FONT
            c.border = _BORDER
            c.alignment = Alignment(vertical="center")
            if col_i == 9:  # VALOR
                c.number_format = '#,##0.00'

    # Freeze panes below headers
    ws.freeze_panes = "A9"

    # Save
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", prefix="bordero_")
    wb.save(tmp.name)
    tmp.close()
    return tmp.name

# ── Endpoints ─────────────────────────────────────────────────────────────────
@app.get("/")
def health():
    return {"status": "ok", "service": "C4 Capital FIDC API",
            "ocr": OCR_OK, "pymupdf": PYMUPDF_OK}

@app.post("/extrair")
async def extrair(files: List[UploadFile] = File(...)):
    resultados = []
    for f in files:
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
        tmp.write(await f.read())
        tmp.close()
        try:
            lista = extrair_pdf(tmp.name)
            for d in lista:
                resultados.append({"ok": True, "arquivo": f.filename, "dados": d})
        except Exception as e:
            resultados.append({"ok": False, "arquivo": f.filename, "erro": str(e)})
        finally:
            os.unlink(tmp.name)
    return {"resultados": resultados}

@app.post("/gerar-bordero")
async def gerar_bordero(files: List[UploadFile] = File(...)):
    todos = []
    for f in files:
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
        tmp.write(await f.read())
        tmp.close()
        try:
            todos.extend(extrair_pdf(tmp.name))
        except Exception:
            pass
        finally:
            os.unlink(tmp.name)

    if not todos:
        return JSONResponse({"erro": "Nenhum dado extraído"}, status_code=422)

    xlsx_path = preencher_bordero(todos)
    return FileResponse(
        xlsx_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="BORDERO_C4.xlsx"
    )
